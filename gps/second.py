import pandas as pd
import numpy as np
import openpyxl
import scipy.stats
import matplotlib.pyplot as plt
from scipy.stats import weibull_max
from scipy.stats import norm
from scipy.stats import expon
from scipy.stats import lognorm
import random
import xlrd
import copy
from sklearn.preprocessing import StandardScaler
from random import choices
import serial as ser
import time
import struct
import subprocess
population = [1, 2]
weights = [0,1]
a=choices(population, weights)
print(a)
no_11 = 0
old_upper_value_wear=0
old_lower_value_wear=0
no_1 = 0
wear_row_no=1
c = 1
no = 1
if (a == [1]):
    print("wear out")
    r = random.randint(26, 30)
    print(r)
if(a == [2]):
    print("breakage")
    r = random.randint(14, 15)
    print("RUL:", r)

for i in range(1,r):
    print(i)
    if (a == [1]):

        b = 1
        loc = "C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx"
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(b)
        sheet1 = wb.sheet_by_index(0)
        wb1 = openpyxl.load_workbook(filename="C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx")
        ws = wb1.worksheets[4]
        sheet.cell_value(0, 0)



    if (a == [2]):

        b = 3
        loc = "C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx"
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(b)
        sheet1 = wb.sheet_by_index(2)
        wb1 = openpyxl.load_workbook(filename="C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx")
        ws = wb1.worksheets[5]
        sheet.cell_value(0, 0)

    e = sheet.row_values(i)
    e1 = copy.deepcopy(e)

    k = []
    for x, value in enumerate(e, 0):
        if value == '':
            k.append(x)

    # print(k)
    for value in sorted(k, reverse=True):
        del e[value]

    # y.remove('')
    print("RMS_Data_points:",e)
    y = np.asarray(e)


    x = np.arange(len(y))
    size = len(y)


    y_df = pd.DataFrame(y, columns=['Data'])
    y_df.describe()

    sc = StandardScaler()
    # print(sc)
    yy = y.reshape(size, 1)
    # print(yy.shape)
    # sc.fit(yy)
    if len(yy) != 1:
        y_std = sc.fit_transform(yy)
        y_std = y_std.flatten()

    else:
        y_std = yy
    del yy
    import warnings

    warnings.filterwarnings("ignore")

    dist_names = ['expon',
                  'norm',
                  'lognorm',
                  'weibull_max']

    # Set up empty lists to stroe results
    chi_square = []
    p_values = []
    percentile_bins = np.linspace(0, 100, 51)
    percentile_cutoffs = np.percentile(y_std, percentile_bins)
    observed_frequency, bins = (np.histogram(y_std, bins=percentile_cutoffs))
    cum_observed_frequency = np.cumsum(observed_frequency)

    for distribution in dist_names:
        # Set up distribution and get fitted distribution parameters
        dist = getattr(scipy.stats, distribution)
        param = dist.fit(y_std)

        p = scipy.stats.kstest(y_std, distribution, args=param)[1]
        p = np.around(p, 5)
        p_values.append(p)

        cdf_fitted = dist.cdf(percentile_cutoffs, *param[:-2], loc=param[-2],
                              scale=param[-1])
        expected_frequency = []
        for bin in range(len(percentile_bins) - 1):
            expected_cdf_area = cdf_fitted[bin + 1] - cdf_fitted[bin]
            expected_frequency.append(expected_cdf_area)

        expected_frequency = np.array(expected_frequency) * size
        cum_expected_frequency = np.cumsum(expected_frequency)
        ss = sum(((cum_expected_frequency - cum_observed_frequency) ** 2) / cum_observed_frequency)
        chi_square.append(ss)

    results = pd.DataFrame()
    results['Distribution'] = dist_names
    results['chi_square'] = chi_square
    results['p_value'] = p_values
    results.sort_values(['chi_square'], inplace=True)

    # Report results

    #print('\nDistributions sorted by goodness of fit:')
    #print('----------------------------------------')
    #print(results)

    number_of_bins = 100
    bin_cutoffs = np.linspace(np.percentile(y, 0), np.percentile(y, 99), number_of_bins)

    h = plt.hist(y, bins=bin_cutoffs, color='0.75')

    number_distributions_to_plot = 1
    dist_names = results['Distribution'].iloc[0:number_distributions_to_plot]

    parameters = []

    for dist_name in dist_names:
        dist = getattr(scipy.stats, dist_name)
        param = dist.fit(y)
        parameters.append(param)

        pdf_fitted = dist.pdf(x, *param[:-2], loc=param[-2], scale=param[-1])
        scale_pdf = np.trapz(h[0], h[1][:-1]) / np.trapz(pdf_fitted, x)
        pdf_fitted *= scale_pdf

    dist_parameters = pd.DataFrame()
    dist_parameters['Distribution'] = (results['Distribution'].iloc[0:number_distributions_to_plot])
    dist_parameters['Distribution parameters'] = parameters

    for index, row in dist_parameters.iterrows():
        print('\n')
        print('Distribution:', row[0])
        print('Parameters:', row[1])
        print('\n')

        # distribution = input("Name of dist: ")
        while True:
            if (row[0] == "weibull_max"):
                rdnum_rms = weibull_max.rvs(row[1][0], row[1][1], row[1][2], 1)
                print("rms_random_no:",rdnum_rms)
            if (row[0] == "norm"):
                rdnum_rms = norm.rvs(row[1][0], row[1][1], 1)
                print("rms_random_no:",rdnum_rms)
            if (row[0] == "lognorm"):
                rdnum_rms = lognorm.rvs(row[1][0], row[1][1], row[1][2], 1)
                print("rms_random_no:",rdnum_rms)
            if (row[0] == "expon"):
                rdnum_rms = expon.rvs(row[1][0], row[1][1], 1)
                print("rms_random_no:",rdnum_rms)

            if (b == 1 and rdnum_rms >= no_1 and rdnum_rms <= max(y)):
                break

            if (b == 3 and rdnum_rms >= no_1 and rdnum_rms <= max(y)):
                break

        no_1 = rdnum_rms

        ws.cell(row=no , column=c).value = rdnum_rms[0]
        no = no + 1
        wb1.save("C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx")

    if ((b==1) and (rdnum_rms >=0.4)):
        break
    if ((b == 3) and (rdnum_rms >= 0.14)):
        break

    e.append(rdnum_rms)
    e.sort()
    idx = e.index(rdnum_rms)
    lower_value = e[idx-1]
    print(lower_value)
    upper_value = e[idx+1]
    print(upper_value)

    for t in range(i,r):
        w = sheet1.row_values(t)
        print(w)
        break

    lower_value_wear = w[e1.index(lower_value)]
    upper_value_wear = w[e1.index(upper_value)]
    if (lower_value_wear > upper_value_wear):
        upper_value_wear= w[e1.index(lower_value)]
        lower_value_wear= w[e1.index(upper_value)]

    if (upper_value_wear < old_upper_value_wear):
        for i1 in e1:
            if (i1>lower_value and upper_value):
                #print([e1.index[i1]])
                upper_value_wear = max(w)
                break
    print("lower_limit_wear:",lower_value_wear)
    print("upper_limit_wear:",upper_value_wear)
    old_upper_value_wear = upper_value_wear
    old_lower_value_wear = lower_value_wear
    #___________________________________________________________________________________________________________________________________________
    if (a == [1]):
        b = 0
        loc = "C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx"
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(b)
        sheet.cell_value(0, 0)
    if (a == [2]):
        b = 2
        loc = "C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx"
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(b)
        sheet.cell_value(0, 0)



    z1 = sheet1.row_values(i)

    k = []
    for x, value in enumerate(z1, 0):
        if value == '':
            k.append(x)

    # print(k)
    for value in sorted(k, reverse=True):
        del z1[value]

    # y.remove('')
    print(z1)
    y = np.asarray(z1)

    x = np.arange(len(y))
    size = len(y)


    y_df = pd.DataFrame(y, columns=['Data'])
    y_df.describe()

    sc = StandardScaler()
    # print(sc)
    yy = y.reshape(size, 1)
    # print(yy.shape)
    # sc.fit(yy)
    if len(yy) != 1:
        y_std = sc.fit_transform(yy)
        y_std = y_std.flatten()

    else:
        y_std = yy
    del yy
    import warnings

    warnings.filterwarnings("ignore")

    dist_names = ['expon',
                  'norm',
                  'lognorm',
                  'weibull_max']

    # Set up empty lists to stroe results
    chi_square = []
    p_values = []
    percentile_bins = np.linspace(0, 100, 51)
    percentile_cutoffs = np.percentile(y_std, percentile_bins)
    observed_frequency, bins = (np.histogram(y_std, bins=percentile_cutoffs))
    cum_observed_frequency = np.cumsum(observed_frequency)

    for distribution in dist_names:
        # Set up distribution and get fitted distribution parameters
        dist = getattr(scipy.stats, distribution)
        param = dist.fit(y_std)

        p = scipy.stats.kstest(y_std, distribution, args=param)[1]
        p = np.around(p, 5)
        p_values.append(p)

        cdf_fitted = dist.cdf(percentile_cutoffs, *param[:-2], loc=param[-2],
                              scale=param[-1])
        expected_frequency = []
        for bin in range(len(percentile_bins) - 1):
            expected_cdf_area = cdf_fitted[bin + 1] - cdf_fitted[bin]
            expected_frequency.append(expected_cdf_area)

        expected_frequency = np.array(expected_frequency) * size
        cum_expected_frequency = np.cumsum(expected_frequency)
        ss = sum(((cum_expected_frequency - cum_observed_frequency) ** 2) / cum_observed_frequency)
        chi_square.append(ss)

    results = pd.DataFrame()
    results['Distribution'] = dist_names
    results['chi_square'] = chi_square
    results['p_value'] = p_values
    results.sort_values(['chi_square'], inplace=True)

    # Report results

    print('\nDistributions sorted by goodness of fit:')
    print('----------------------------------------')
    print(results)

    number_of_bins = 100
    bin_cutoffs = np.linspace(np.percentile(y, 0), np.percentile(y, 99), number_of_bins)

    h = plt.hist(y, bins=bin_cutoffs, color='0.75')

    number_distributions_to_plot = 1
    dist_names = results['Distribution'].iloc[0:number_distributions_to_plot]

    parameters = []

    for dist_name in dist_names:
        dist = getattr(scipy.stats, dist_name)
        param = dist.fit(y)
        parameters.append(param)

        pdf_fitted = dist.pdf(x, *param[:-2], loc=param[-2], scale=param[-1])
        scale_pdf = np.trapz(h[0], h[1][:-1]) / np.trapz(pdf_fitted, x)
        pdf_fitted *= scale_pdf

    dist_parameters = pd.DataFrame()
    dist_parameters['Distribution'] = (results['Distribution'].iloc[0:number_distributions_to_plot])
    dist_parameters['Distribution parameters'] = parameters

    for index, row in dist_parameters.iterrows():
        print('\n')
        print('Distribution:', row[0])
        print('Parameters:', row[1])
        print('\n')

        # distribution = input("Name of dist: ")
        while True:
            if (row[0] == "weibull_max"):
                rdnum_wear = weibull_max.rvs(row[1][0], row[1][1], row[1][2], 1)
                print("Wear:",rdnum_wear)
            if (row[0] == "norm"):
                rdnum_wear = norm.rvs(row[1][0], row[1][1], 1)
                print("Wear:",rdnum_wear)
            if (row[0] == "lognorm"):
                rdnum_wear = lognorm.rvs(row[1][0], row[1][1], row[1][2], 1)
                print("Wear:",rdnum_wear)
            if (row[0] == "expon"):
                rdnum_wear = expon.rvs(row[1][0], row[1][1], 1)
                print("Wear:",rdnum_wear)


            if (b == 0 and rdnum_wear >= no_11 and lower_value_wear <=rdnum_wear <= upper_value_wear):
                break

            if (b == 2 and rdnum_wear >= no_11 and lower_value_wear <=rdnum_wear <= upper_value_wear):
                break

        no_11 = rdnum_wear
        ws.cell(row= wear_row_no  , column=c+1).value = rdnum_wear[0]
        wear_row_no = wear_row_no + 1
        wb1.save("C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx")
        #plot this wera values in third graph
    if ((b == 0) and (rdnum_wear >= 0.8)):
        break
    if ((b == 2) and (rdnum_wear >= 0.75)):
        break


#_______________________________________________________________________________________________________________________________________________

loc = ("C:\\Users\\Rushabh Kadam\\Desktop\\WEAR AND RMS 2.xlsx")

wb = xlrd.open_workbook(loc)
if (a == [1]):
    sheet = wb.sheet_by_index(4)
if (a == [2]):
    sheet = wb.sheet_by_index(5)
sheet.cell_value(0, 0)
rms_gen = []
rms_inc = [1]

for i in range(sheet.nrows):
    rms_gen.append(sheet.cell_value(i, 0))
#print(rms_gen)

for i in range(len(rms_gen)-1):
    rms_inc.append(((rms_gen[i+1] - rms_gen[i])/rms_gen[i])+1)
#print(rms_inc)

loc1 = ("C:\\Users\\Rushabh Kadam\\Desktop\\Data bank.xlsx")

wb11 = xlrd.open_workbook(loc1)
sheet11 = wb11.sheet_by_index(0)
rms_databank = []
analog_write =[]
for i in range(sheet11.nrows):
    rms_databank.append(sheet11.cell_value(i, 1))
    analog_write.append(sheet11.cell_value(i,0))
#print(rms_databank)

motor_rms = []
first_value = rms_databank[0]
for i in range(len(rms_inc)):
    motor_rms.append(first_value*rms_inc[i])
    first_value = motor_rms[i]
#print(motor_rms)

def find_nearest(array, value):
    array = np.asarray(array)
    idx = (np.abs(array - value)).argmin()
    return array[idx]

motor_rms_near_value =[]
for i in range(len(motor_rms)):
    motor_rms_near_value.append(find_nearest(rms_databank,motor_rms[i] ))
#print(motor_rms_near_value)

analog_write_near_value=[]
for i in motor_rms_near_value:
    analog_write_near_value.append(analog_write[rms_databank.index(i)])
print(analog_write_near_value)
analog_write_near_value = list(map(int,analog_write_near_value))
print(analog_write_near_value)
print(len(analog_write_near_value))



ser = ser.Serial('COM3',9600)
subprocess.Popen(["python", "first.py", str(len(analog_write_near_value))], shell=True)
time.sleep(6)

for a1 in analog_write_near_value:
    ser.write(struct.pack('i',a1))
    time.sleep(5)

ser.close()


#_____________________________________________________________________________________________________________________________________________







#_______________________________________________________________________________________________________________________________________________














